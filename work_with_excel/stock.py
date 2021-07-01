from openpyxl import Workbook, load_workbook
import os
import datetime

folder = "./records/"
companyName = "EXPO.N0000.xlsx"
filename = folder + companyName

fileExists = os.path.exists(filename)

def insertToSecondRow(a,b,c,d,e,f,g,h,i,j,k,l):
    ws['A2'] = datetime.date.today()
    ws['B2'] = a
    ws['C2'] = b
    ws['D2'] = c
    ws['E2'] = d
    ws['F2'] = e
    ws['G2'] = f
    ws['H2'] = g
    ws['I2'] = h
    ws['J2'] = i
    ws['K2'] = j
    ws['J2'] = k
    ws['M2'] = l

if (fileExists):
    wb = load_workbook(filename = filename)
    ws = wb.active
    ws.insert_rows(2)
    ws['A2']=datetime.date.today()
    ws['B2'] = 556
    ws['C2'] = 455.6

    wb.save(filename)

else:
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws.append(['Date','Company Name','Profit'])
    wb.save(filename)